import streamlit as st
import openpyxl
import json
import re
import time
from openpyxl import Workbook
from collections import defaultdict, Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from copy import copy
from tencentcloud.common import credential
from tencentcloud.tmt.v20180321 import tmt_client, models
from deep_translator import GoogleTranslator
from io import BytesIO
import pandas as pd

# ===================== 初始化 =====================
if "translated" not in st.session_state:
    st.session_state.translated = False
if "editor_keywords" not in st.session_state:
    st.session_state.editor_keywords = {}
if "last_range" not in st.session_state:
    st.session_state.last_range = (None, None)
if "full_doc_mode" not in st.session_state:
    st.session_state.full_doc_mode = False
if "SecretId" not in st.session_state:
    st.session_state.SecretId = ""
if "SecretKey" not in st.session_state:
    st.session_state.SecretKey = ""
if "translating" not in st.session_state:
    st.session_state.translating = False

st.set_page_config(page_title="期刊翻译与关键词提取工具", layout="wide")
st.title("📚 期刊翻译与关键词提取工具")

# ===================== API 密钥配置 =====================
st.sidebar.header("🔐 腾讯翻译 API 配置")
secret_id_input = st.sidebar.text_input("SecretId", type="password", value=st.session_state.SecretId)
secret_key_input = st.sidebar.text_input("SecretKey", type="password", value=st.session_state.SecretKey)
remember = st.sidebar.checkbox("记住密码")

if st.sidebar.button("✅ 载入密钥"):
    if secret_id_input and secret_key_input:
        st.session_state.SecretId = secret_id_input
        st.session_state.SecretKey = secret_key_input
        st.success("✅ 密钥载入成功")
        if remember:
            with open("config.json", "w", encoding="utf-8") as f:
                json.dump({"SecretId": secret_id_input, "SecretKey": secret_key_input}, f, ensure_ascii=False, indent=2)
            st.info("🔒 密钥已保存到 config.json")
    else:
        st.warning("⚠️ 请填写完整的密钥信息")

@st.cache_resource(show_spinner=False)
def get_translation_client(secret_id, secret_key):
    cred = credential.Credential(secret_id, secret_key)
    return tmt_client.TmtClient(cred, "ap-beijing"), GoogleTranslator(source='en', target='zh-CN')

client, translator_backup = None, None
if st.session_state.SecretId and st.session_state.SecretKey:
    try:
        client, translator_backup = get_translation_client(st.session_state.SecretId, st.session_state.SecretKey)
    except Exception as e:
        st.error(f"⚠️ 初始化翻译客户端失败：{e}")

# ===================== 文件上传 & 范围 =====================
uploaded_file = st.file_uploader("📄 上传 Excel 文件", type=["xlsx"])
st.session_state.full_doc_mode = st.checkbox("📘 读取全文档进行翻译")

if not st.session_state.full_doc_mode:
    start_row = st.number_input("起始行", min_value=2, value=2, key="start")
    end_row = st.number_input("结束行", min_value=start_row, value=start_row + 10, key="end")
else:
    start_row, end_row = None, None

# ===================== 翻译函数 =====================
TERMS_MAP = {
    "journal": "期刊", "editor": "编辑", "publication": "出版物",
    "manuscript": "稿件", "article": "文章", "tumor": "肿瘤",
    "cell": "细胞", "protein": "蛋白质", "COVID-19": "新冠病毒"
}
TRANSLATION_CACHE = {}

def clean_text(text):
    return str(text).replace('\n', ' ').replace('\r', ' ').strip()

def normalize_translation(text):
    for en, zh in TERMS_MAP.items():
        text = text.replace(en, zh)
    return text

def translate_text(text):
    if not text or not re.search(r'[a-zA-Z]', text):
        return text
    cleaned = clean_text(text)
    if cleaned in TRANSLATION_CACHE:
        return TRANSLATION_CACHE[cleaned]
    try:
        req = models.TextTranslateRequest()
        params = {"SourceText": cleaned, "Source": "en", "Target": "zh", "ProjectId": 0}
        req.from_json_string(json.dumps(params))
        resp = client.TextTranslate(req)
        result = normalize_translation(resp.TargetText)
    except:
        try:
            result = normalize_translation(translator_backup.translate(cleaned))
        except:
            result = f"[失败]{text}"
    TRANSLATION_CACHE[cleaned] = result
    return result

def batch_translate(texts, label="翻译中"):
    results = [None] * len(texts)
    total = len(texts)
    start_time = time.time()
    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = {executor.submit(translate_text, t): i for i, t in enumerate(texts)}
        completed = 0
        progress_bar = st.progress(0, text=f"{label}：0/{total}")
        for future in as_completed(futures):
            i = futures[future]
            results[i] = future.result()
            completed += 1
            elapsed = time.time() - start_time
            rate = completed / elapsed if elapsed > 0 else 0
            remaining = total - completed
            eta = remaining / rate if rate > 0 else 0
            progress_bar.progress(completed / total, text=f"{label}：{completed}/{total}，预计剩余 {eta:.1f} 秒")
    return results
# ✅ 衔接前面已写的部分，这里继续添加核心逻辑

# Excel 数据处理与输出（续）
if uploaded_file:
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active
    max_row = ws.max_row

    if st.session_state.full_doc_mode:
        start_row = 2
        end_row = max_row

    disable_btn = st.session_state.get("translating", False) or (
        st.session_state.translated and st.session_state.last_range == (start_row, end_row)
    )
    translate_clicked = st.button("🚀 开始翻译", disabled=disable_btn)

    if translate_clicked:
        st.session_state.translating = True

        # 抽取标题、关键词、作者
        titles = [ws.cell(row=r, column=3).value or "" for r in range(start_row, end_row + 1)]
        keywords = [ws.cell(row=r, column=6).value or "" for r in range(start_row, end_row + 1)]
        editors = [ws.cell(row=r, column=1).value or "匿名" for r in range(start_row, end_row + 1)]

        # 翻译
        st.subheader("⏳ 正在翻译标题")
        translated_titles = batch_translate(titles, label="标题翻译")

        st.subheader("⏳ 正在翻译关键词")
        translated_keywords = batch_translate(keywords, label="关键词翻译")

        # 写入新文件
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = ws.title

        row_map = list(range(start_row, end_row + 1))
        editor_keywords = defaultdict(list)

        for idx, row_num in enumerate(row_map):
            for col in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=row_num, column=col)
                dst_cell = new_ws.cell(row=2 * idx + 1, column=col, value=src_cell.value)
                dst_cell.font = copy(src_cell.font)
                dst_cell.alignment = copy(src_cell.alignment)

            new_ws.cell(row=2 * idx + 2, column=3, value=f"标题：{translated_titles[idx]}")
            new_ws.cell(row=2 * idx + 2, column=6, value=f"关键词：{translated_keywords[idx]}")

            # 提取关键词并归类
            editor = editors[idx]
            raw_kw = translated_keywords[idx].replace("关键词：", "").strip()
            for kw in re.split(r"[；;]", raw_kw):
                cleaned = kw.strip()
                if cleaned:
                    editor_keywords[editor].append(cleaned)

        # 去重关键词
        editor_keywords = {
            editor: dict(Counter(set(kws))) for editor, kws in editor_keywords.items()
        }

        st.session_state.editor_keywords = editor_keywords
        st.session_state.translated = True
        st.session_state.last_range = (start_row, end_row)
        st.session_state.translating = False

        output = BytesIO()
        new_wb.save(output)
        output.seek(0)
        st.download_button("📥 下载翻译结果 Excel", data=output, file_name="翻译结果.xlsx")

# 展示关键词界面
if st.session_state.translated and st.session_state.editor_keywords:
    st.subheader("📌 按编委查看关键词")
    selected = st.selectbox("选择编委姓名", sorted(k for k in st.session_state.editor_keywords if isinstance(k, str)))
    if selected:
        st.markdown(f"#### ✨ {selected} 的关键词：")
        kw_freq = st.session_state.editor_keywords[selected]

        cols = st.columns(4)
        for i, (kw, count) in enumerate(sorted(kw_freq.items(), key=lambda x: -x[1])):
            with cols[i % 4]:
                st.markdown(f"<span style='display:inline-block;padding:6px 12px;border-radius:8px;background:#f0f0f0;'>{kw} ({count}次)</span>", unsafe_allow_html=True)

        df_export = [f"{kw} ({count}次)" for kw, count in sorted(kw_freq.items(), key=lambda x: -x[1])]
        df_bytes = BytesIO()
        df_str = "\n".join(df_export)
        df_bytes.write(df_str.encode("utf-8"))
        df_bytes.seek(0)
        st.download_button("⬇️ 下载该编委关键词（TXT）", df_bytes, file_name=f"{selected}_关键词.txt")
